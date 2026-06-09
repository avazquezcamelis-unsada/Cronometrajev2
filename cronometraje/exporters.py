from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

from .database import Database
from .utils import calculate_age, format_date, format_dni, iso_to_display, seconds_to_time

def get_user_exports_dir() -> Path:
    """Carpeta segura para exportar Excel/PDF, funcione como .py o como .exe instalado."""
    if os.name == "nt":
        docs = Path.home() / "Documents"
        if docs.exists():
            return docs / "Cronometraje" / "exports"
        root = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
        if root:
            return Path(root) / "Cronometraje" / "exports"
    return Path.home() / "Cronometraje" / "exports"


EXPORTS_DIR = get_user_exports_dir()
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def safe_name(value: str) -> str:
    clean = "".join(ch if ch.isalnum() else "_" for ch in value.strip())
    return clean.strip("_") or "carrera"


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def build_podiums(db: Database, id_carrera: int) -> dict[str, list[dict[str, Any]]]:
    rows = db.list_results(id_carrera)
    podiums: dict[str, list[dict[str, Any]]] = {}

    for distancia in ["6K", "12K", "18K"]:
        for sexo in ["M", "F"]:
            key = f"General {distancia} {sexo}"
            items = [r for r in rows if r["distancia"] == distancia and r["sexo"] == sexo]
            podiums[key] = sorted(items, key=lambda x: x["tiempo_llegada"])[:5]

    categories = sorted({r["categoria"] for r in rows})
    for distancia in ["6K", "12K", "18K"]:
        for sexo in ["M", "F"]:
            for categoria in categories:
                items = [
                    r for r in rows
                    if r["distancia"] == distancia and r["sexo"] == sexo and r["categoria"] == categoria
                ]
                if items:
                    key = f"Categoria {distancia} {sexo} {categoria}"
                    podiums[key] = sorted(items, key=lambda x: x["tiempo_llegada"])[:3]
    return podiums


def export_excel(db: Database, id_carrera: int) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Falta instalar openpyxl. Ejecutá: pip install -r requirements.txt") from exc

    race = db.get_race(id_carrera)
    if not race:
        raise ValueError("No se encontró la carrera")

    filename = f"cronometraje_{safe_name(race['nombre'])}_{timestamp()}.xlsx"
    path = EXPORTS_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Corredores"

    headers = ["Nº", "DNI", "Apellido", "Nombre", "Sexo", "Ciudad", "Fecha nac.", "Edad", "Team", "KM", "Talle", "Categoría", "Estado"]
    ws.append(headers)
    for row in db.list_enrollments(id_carrera):
        edad = calculate_age(row["fecha_nacimiento"], row["fecha_carrera"])
        ws.append([
            row["numero"], format_dni(row["dni"]), row["apellido"], row["nombre"], row["sexo"], row["ciudad"],
            format_date(row["fecha_nacimiento"]), edad, row["team"], row["distancia"], row["talle"], row["categoria"], row["estado"],
        ])

    ws2 = wb.create_sheet("Llegadas")
    ws2.append(["Tiempo", "Hora registro", "Nº", "Distancia", "Categoría", "Sexo", "Apellido", "Nombre", "Ciudad", "Tipo", "Observación"])
    for row in db.list_arrivals(id_carrera):
        ws2.append([
            seconds_to_time(row["tiempo_llegada"]), iso_to_display(row["hora_registro"]), row["numero"], row["distancia"],
            row["categoria"], row["sexo"], row["apellido"], row["nombre"], row["ciudad"], row["tipo_registro"], row["observacion"],
        ])

    ws3 = wb.create_sheet("Resultados")
    ws3.append(["Posición", "Tiempo", "Nº", "Distancia", "Categoría", "Sexo", "Apellido", "Nombre", "Ciudad"])
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in db.list_results(id_carrera):
        grouped.setdefault((row["distancia"], row["sexo"]), []).append(row)
    for key, items in sorted(grouped.items()):
        for pos, row in enumerate(sorted(items, key=lambda x: x["tiempo_llegada"]), start=1):
            ws3.append([pos, seconds_to_time(row["tiempo_llegada"]), row["numero"], row["distancia"], row["categoria"], row["sexo"], row["apellido"], row["nombre"], row["ciudad"]])

    ws4 = wb.create_sheet("Sin tiempo")
    ws4.append(["Nº", "DNI", "Apellido", "Nombre", "Sexo", "Distancia", "Categoría", "Estado"])
    for row in db.list_without_time(id_carrera):
        ws4.append([row["numero"], format_dni(row["dni"]), row["apellido"], row["nombre"], row["sexo"], row["distancia"], row["categoria"], row["estado"]])

    ws5 = wb.create_sheet("Podios")
    ws5.append(["Podio", "Puesto", "Tiempo", "Nº", "Apellido", "Nombre", "Sexo", "Distancia", "Categoría", "Ciudad"])
    for title, items in build_podiums(db, id_carrera).items():
        for pos, row in enumerate(items, start=1):
            ws5.append([title, pos, seconds_to_time(row["tiempo_llegada"]), row["numero"], row["apellido"], row["nombre"], row["sexo"], row["distancia"], row["categoria"], row["ciudad"]])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), 35)

    wb.save(path)
    return path


def export_pdf(db: Database, id_carrera: int, report_type: str = "todos") -> Path:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    except ImportError as exc:
        raise RuntimeError("Falta instalar reportlab. Ejecutá: pip install -r requirements.txt") from exc

    race = db.get_race(id_carrera)
    if not race:
        raise ValueError("No se encontró la carrera")

    filename = f"reporte_{report_type}_{safe_name(race['nombre'])}_{timestamp()}.pdf"
    path = EXPORTS_DIR / filename
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    def title(text: str):
        story.append(Paragraph(text, styles["Title"]))
        story.append(Paragraph(f"{race['nombre']} - {format_date(race['fecha'])} - {race['lugar'] or ''}", styles["Normal"]))
        story.append(Spacer(1, 12))

    def add_table(headers: list[str], rows: list[list[Any]]):
        data = [headers] + rows
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 12))

    if report_type in ("todos", "general"):
        title("Resultado general por distancia")
        rows = []
        grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in db.list_results(id_carrera):
            grouped.setdefault((row["distancia"], row["sexo"]), []).append(row)
        for key, items in sorted(grouped.items()):
            for pos, row in enumerate(sorted(items, key=lambda x: x["tiempo_llegada"]), start=1):
                rows.append([pos, seconds_to_time(row["tiempo_llegada"]), row["numero"], row["distancia"], row["sexo"], row["categoria"], row["apellido"], row["nombre"], row["ciudad"]])
        add_table(["Pos", "Tiempo", "Nº", "KM", "Sexo", "Categoría", "Apellido", "Nombre", "Ciudad"], rows or [["Sin datos", "", "", "", "", "", "", "", ""]])
        if report_type == "todos":
            story.append(PageBreak())

    if report_type in ("todos", "podios_generales", "podios_categorias"):
        title("Podios")
        rows = []
        for podio, items in build_podiums(db, id_carrera).items():
            if report_type == "podios_generales" and not podio.startswith("General"):
                continue
            if report_type == "podios_categorias" and not podio.startswith("Categoria"):
                continue
            for pos, row in enumerate(items, start=1):
                rows.append([podio, pos, seconds_to_time(row["tiempo_llegada"]), row["numero"], row["apellido"], row["nombre"], row["ciudad"]])
        add_table(["Podio", "Puesto", "Tiempo", "Nº", "Apellido", "Nombre", "Ciudad"], rows or [["Sin datos", "", "", "", "", "", ""]])
        if report_type == "todos":
            story.append(PageBreak())

    if report_type in ("todos", "llegadas"):
        title("Listado completo de llegadas")
        rows = []
        for row in db.list_arrivals(id_carrera):
            rows.append([seconds_to_time(row["tiempo_llegada"]), row["numero"], row["distancia"], row["categoria"], row["sexo"], row["apellido"], row["nombre"], row["ciudad"], row["tipo_registro"]])
        add_table(["Tiempo", "Nº", "KM", "Categoría", "Sexo", "Apellido", "Nombre", "Ciudad", "Tipo"], rows or [["Sin datos", "", "", "", "", "", "", "", ""]])
        if report_type == "todos":
            story.append(PageBreak())

    if report_type in ("todos", "sin_tiempo"):
        title("Corredores sin tiempo")
        rows = []
        for row in db.list_without_time(id_carrera):
            rows.append([row["numero"], format_dni(row["dni"]), row["apellido"], row["nombre"], row["sexo"], row["distancia"], row["categoria"], row["estado"]])
        add_table(["Nº", "DNI", "Apellido", "Nombre", "Sexo", "KM", "Categoría", "Estado"], rows or [["Sin datos", "", "", "", "", "", "", ""]])

    doc.build(story)
    return path
